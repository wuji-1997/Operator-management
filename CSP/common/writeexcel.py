# -- coding: utf-8 --
import logging
from common.my_log import Log
import openpyxl

#excel文档的基本定义
'''
1) 工作薄(workbook):一个EXCEL文件就称为一个工作薄,一个工作薄中可以包含若干张工作表。
2) 工作表(sheet):工作薄中的每一张表格称为工作表,每张工作表都有一个标签,默认为sheet1\sheet2\sheet3来命名，(一个工作 薄默认为由3个工作表组成)
3) 活动表(active sheet)：指当前正在操作的工作表
4) 行(row): 工作表中的每一行行首数字(1、2、3、)称为行标题;一张工作表最多有65536行
5) 列(column): 列标题:工作表中每一列列首的字母(A、B、C)称为列标题;一张工作表最多有256列
6)单元格(cell): 工作表的每一个格称为单元格

'''

#1.创建对象
'''
打开或者创建一个Excel需要创建一个Workbook对象
获取一个表则需要先创建一个Workbook对象，然后使用该对象的方法来得到一个Worksheet对象
如果要获取表中的数据，那么得到Worksheet对象以后再从中获取代表单元格的Cell对象
Workbook对象
一个Workbook对象代表一个Excel文档，因此在操作Excel之前，都应该先创建一个Workbook对象。对于创建一个新的Excel文档，
直接进行Workbook类的调用即可，对于一个已经存在的Excel文档，可以使用openpyxl模块的load_workbook函数进行读取，该函数包涵多个参数，
但只有filename参数为必传参数。filename 是一个文件名，也可以是一个打开的文件对象
'''
#2.Workbook对象属性（工作簿操作）
'''
sheetnames：获取工作簿中的表（列表）
active：获取当前活跃的Worksheet
worksheets：以列表的形式返回所有的Worksheet(表格)
read_only：判断是否以read_only模式打开Excel文档
encoding：获取文档的字符集编码
properties：获取文档的元数据，如标题，创建者，创建日期等
'''
#3.Worksheet，Cell对象（工作表操作，单元格）
'''
Worksheet:
title：表格的标题
max_row：表格的最大行
min_row：表格的最小行
max_column：表格的最大列
min_column：表格的最小列
rows：按行获取单元格(Cell对象) - 生成器
columns：按列获取单元格(Cell对象) - 生成器
values：按行获取表格的内容(数据) - 生成器
'''

excel_log = Log(__name__,file=logging.INFO,cmd=logging.INFO)

class Write_Excel():

    def __init__(self,filepath,number):
        """

        :param filename:
        """

        self.filepath = filepath
        try:
            self.wb = openpyxl.load_workbook(filename=filepath)
            self.worksheet = self.wb.worksheets[number]
            sheet_name = self.worksheet.title
        except Exception:
            excel_log.csp_log.exception('打开excel文件失败或sheet读取错误')
            raise
        else:
            excel_log.csp_log.info(f'打开excel文件且读取sheet-----{sheet_name}------成功')


    def read_worksheet(self,sheetrow,sheetcolum):
        """

        :param number:
        :param sheetrow:
        :param sheetcolum:
        :return:
        """
        try:
            data = self.worksheet.cell(sheetrow,sheetcolum).value
        except Exception:
            excel_log.csp_log.exception('获取数据失败')
            raise
        else:
            excel_log.csp_log.exception(f'获取单元格数据成功------{data}')
            return data
    def read_row(self,sheetrow):
        """
        读取一整行数据
        :param sheetrow:
        :return:
        """
        try:
            rows = self.worksheet.max_column
            data = []
            for i in range(1, rows + 1):
                value = self.worksheet.cell(row=sheetrow, column=i).value
                data.append(value)

        except Exception:
            excel_log.csp_log.exception('获取一整行数据失败')
            raise
        else:
            excel_log.csp_log.exception(f'获取一整行数据成功')
            return data

    def read_column(self,sheetcolumn):
        """
        读取一整列数据
        :param sheetcolumn:
        :return:
        """

        try:
            rows = self.worksheet.max_row
            data = []
            for i in range(1,rows+1):
                value = self.worksheet.cell(row=i,column=sheetcolumn).value
                data.append(value)

        except Exception:
            excel_log.csp_log.exception('获取一整列数据失败')
            raise
        else:
            excel_log.csp_log.exception(f'获取一整列数据成功')
            return data


    def update_data(self,row,clow,update):
        """

        :param row:
        :param clow:
        :param update:
        :return:
        """

        try:
            self.worksheet.cell(row,clow,value=update)
        except Exception:
            excel_log.csp_log.exception(f'更新或插入数据---{update}--失败')
            raise
        else:
            excel_log.csp_log.exception(f'更新或插入数据---{update}--成功')
            self.wb.save(filename=self.filepath)

if __name__=="__main__":
    from config.Conf import *
    import time
    import locale
    #test = Write_Excel(filepath=test_data+r'\test_chatbot_data.xlsx',number=3)
    #value =test.read_worksheet(1,1)
    #value2 = test.read_worksheet(1,2)
    #test.update_data(1,1,'序号')
    #valeu =test.read_column(4)[1]
    #number = test.read_column(5)[19]
    #value = test.read_row(5)
    locale.setlocale(locale.LC_CTYPE, 'chinese')
    now = time.strftime('%Y-%m-%d %H:%M:%S')
    test=Write_Excel(filepath=test_data+r'\test_customer_data.xlsx',number=4)
    value = test.update_data(2,9,f'{now}案例执行通过')
    print(value)




